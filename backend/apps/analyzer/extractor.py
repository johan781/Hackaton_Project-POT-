"""
POT File Parser — Sin IA.

Formatos PDF soportados
───────────────────────
Formato A  "Pivijay / reportes por punto"
  Varias tablas por página. Cada tabla de ensayo comienza con
  "ENSAYO DE CARGA HORIZONTAL", "ENSAYO DE CARGA VERTICAL A TRACION", etc.
  Columnas: Carga real (kg) | Carga real (kN) | Desplaz. (mm)
  Metadata: tabla con PROYECTO / POT / UBICACIÓN / PERFIL / FECHA / CORDENADA

Formato B  "Enexa / formato de campo"
  Una tabla grande por página con filas de escalones (0%, 25%, …)
  y columnas agrupadas por tipo de ensayo.

Formato C  "San Martín / informe generado"
  Texto limpio con marcadores "Punto Analizado:" y "Ensayo: X".

Formato D  "Enexa FORMATO PARA PRUEBA PULL OUT TEST (texto espaciado)"
  Una página por punto. Encabezado con texto espaciado "F O R M A T O P A R A …".
  Filas de escalón con 3 grupos de columnas (Compresión, Tracción, Lateral).
  Punto ID al final: "P 0 T A -1 - S A T" → "POT-A-1-SAT".

XLSX: pares de columnas (Desplazamiento / Fuerza) con etiquetas.
TXT : regex.
"""
from __future__ import annotations

import io
import re
from datetime import date

import openpyxl

KGF_TO_KN = 0.00980665

TIPO_LABELS = {
    "tension_vertical":    "Prueba de Tensión Vertical",
    "compresion_vertical": "Prueba de Compresión Vertical",
    "carga_lateral":       "Prueba de Carga Lateral",
}

TIPO_KEYWORDS: dict[str, list[str]] = {
    "compresion_vertical": ["compresión", "compresion", "compression", "compres"],
    "tension_vertical":    ["tensión", "tension", "tracción", "traccion", "tract"],
    "carga_lateral":       ["lateral", "horizontal", "cortante", "shear"],
}
TIPO_SEQUENCE = ["tension_vertical", "compresion_vertical", "carga_lateral"]


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _detect_tipo(text: str) -> str | None:
    t = text.lower().strip()
    for tipo, kws in TIPO_KEYWORDS.items():
        if any(kw in t for kw in kws):
            return tipo
    return None


def _num(val) -> float | None:
    if val is None:
        return None
    s = re.sub(r"[^\d\.\-\+]", "", str(val).replace(",", ".").strip())
    try:
        v = float(s)
        return v
    except ValueError:
        return None


def _is_missing(val) -> bool:
    return val is None or str(val).strip() in ("", "-", "—", "–", "N/A", "na", "n/a", "NA")


def _clean(val) -> str:
    return str(val or "").strip()


def _empty_punto(pid: str) -> dict:
    return {
        "punto_id": pid,
        "profundidad_m": None, "tipo_perfil": None,
        "coordenadas": None,   "fecha_ensayo": None,
        "observaciones": None, "ensayos": [],
    }


def _build_ensayo(tipo: str, pairs: list[tuple]) -> dict | None:
    """Construye ensayo desde pares (desplazamiento_mm, fuerza_kg)."""
    pts = []
    for disp, fuerza in pairs:
        d = _num(disp)
        f = _num(fuerza)
        if d is not None and f is not None and f > 0:
            pts.append({"desplazamiento_mm": d, "fuerza_kg": f})
    if not pts:
        return None
    return {
        "tipo": tipo,
        "nombre": TIPO_LABELS.get(tipo, tipo),
        "carga_maxima_kgf":         max(p["fuerza_kg"]         for p in pts),
        "desplazamiento_maximo_mm": max(p["desplazamiento_mm"] for p in pts),
        "puntos": pts,
    }


# ─── Formato A: tablas individuales por ensayo (Pivijay) ──────────────────────

# Palabras clave que indican el inicio de una tabla de ensayo
_ENSAYO_TITLE_RE = re.compile(
    r"ensayo\s+de\s+carga\s+(horizontal|vertical|lateral|cortante|tracci|compresi)",
    re.IGNORECASE,
)

# Indicadores de filas de carga vs descarga (filtrar descarga)
_DESCARGA_RE = re.compile(r"^\s*0[,\.]?0?\s*$")


def _decode_repeated(text: str) -> str:
    """'PPPPOOOOTTT' → 'POT'  (caracteres repetidos en múltiplos de 4 en fuentes embebidas).
    Cada grupo de 4 repeticiones representa 1 carácter original:
    'PPPPVVVV 11111111' → 'PV 11'  (4→1, 4→1, 8→2 unos)
    """
    if not text:
        return text
    result, i = [], 0
    chars = list(text)
    while i < len(chars):
        c, count = chars[i], 1
        while i + count < len(chars) and chars[i + count] == c:
            count += 1
        # Si está repetido en múltiplos de 4, decodificar como count//4 copias
        if count >= 4 and count % 4 == 0:
            result.append(c * (count // 4))
        else:
            result.append(c * count)
        i += count
    return "".join(result)


def _find_col(headers: list[str], *keywords) -> int | None:
    """Devuelve el índice de la primera columna cuyo header contiene alguna keyword."""
    for ki in keywords:
        for j, h in enumerate(headers):
            if ki.lower() in h.lower():
                return j
    return None


def _parse_ensayo_table_a(table: list[list]) -> tuple[str | None, list[dict]]:
    """
    Parsea una tabla de tipo Pivijay.
    Devuelve (tipo, lista_de_puntos).
    """
    if not table:
        return None, []

    # Fila 0: título de la tabla
    title = " ".join(_decode_repeated(_clean(c)) for c in table[0] if c)
    tipo = _detect_tipo(title)
    if tipo is None:
        return None, []

    # Buscar fila de cabeceras (primera fila con "Carga" o "Desplaz")
    header_idx = None
    for i, row in enumerate(table[1:], 1):
        row_text = " ".join(_clean(c) for c in row).lower()
        if "carga" in row_text or "desplaz" in row_text:
            header_idx = i
            break
    if header_idx is None:
        return tipo, []

    raw_headers = [_clean(c) for c in table[header_idx]]

    # Detectar columnas de interés
    # Prioridad: Carga real (kg) > Carga nominal (kg)
    col_fuerza_kg = (
        _find_col(raw_headers, "carga real\n(kg)", "carga real (kg)", "real\n(kg)", "real (kg)")
        or _find_col(raw_headers, "carga nominal\n(kg)", "carga nominal (kg)", "nominal\n(kg)")
    )
    # Carga real (kN) como fallback
    col_fuerza_kn = _find_col(raw_headers, "carga real\n(kn)", "carga real (kn)",
                               "real\n(kn)", "real (kn)", "carga real")

    col_disp = _find_col(raw_headers, "desplaz.\n(mm)", "desplaz. (mm)",
                          "desplazamiento\n(mm)", "desplazamiento (mm)", "desplaz")

    if col_disp is None:
        return tipo, []

    # Seleccionar columna de fuerza (kg preferido sobre kN)
    if col_fuerza_kg is not None:
        col_fuerza = col_fuerza_kg
        to_kg = 1.0          # ya está en kg
    elif col_fuerza_kn is not None:
        col_fuerza = col_fuerza_kn
        to_kg = 1.0 / KGF_TO_KN  # kN → kgf
    else:
        # Último recurso: buscar cualquier columna numérica > desplazamiento
        return tipo, []

    # Extraer puntos (solo fase de carga: fuerza > umbral)
    pairs: list[tuple] = []
    for row in table[header_idx + 1:]:
        if len(row) <= max(col_fuerza, col_disp):
            continue
        fv = row[col_fuerza]
        dv = row[col_disp]
        if _is_missing(fv) or _is_missing(dv):
            continue
        f = _num(fv)
        d = _num(dv)
        if f is None or d is None or f <= 0:
            continue
        pairs.append((d, f * to_kg))

    # Deduplicar: para cada nivel de carga similar, quedarse con el
    # punto de mayor desplazamiento (fase de mantenimiento, no inicial)
    if pairs:
        pairs_sorted = sorted(pairs, key=lambda p: p[1])
        deduped: list[tuple] = []
        for disp, fuerza in pairs_sorted:
            # Si hay un punto con fuerza muy similar (±3%), actualizar desplazamiento
            if deduped and abs(fuerza - deduped[-1][1]) / max(deduped[-1][1], 1) < 0.03:
                # Quedarse con el mayor desplazamiento para ese nivel de carga
                if disp > deduped[-1][0]:
                    deduped[-1] = (disp, fuerza)
            else:
                deduped.append((disp, fuerza))
        pairs = deduped

    pts = []
    for disp, fuerza in pairs:
        pts.append({"desplazamiento_mm": disp, "fuerza_kg": fuerza})

    return tipo, pts


def _parse_metadata_table_a(table: list[list]) -> dict:
    """Extrae metadata de la tabla de cabecera del punto (Tabla 1 en Pivijay)."""
    meta: dict = {}
    lat: str | None = None
    lon: str | None = None

    for row in table:
        cells = [_decode_repeated(_clean(c)) for c in row]
        row_text = " ".join(cells).lower()

        # Punto ID: buscar celda exactamente "POT" y tomar el valor a su derecha
        if not meta.get("punto_id"):
            for j, c in enumerate(cells):
                if c.strip().upper() == "POT":
                    for k in range(j + 1, len(cells)):
                        v = cells[k].strip()
                        if v and v.upper() != "POT":
                            meta["punto_id"] = "POT-" + v.replace(" ", "-")
                            break
                    break

        if not meta.get("proyecto") and "proyecto" in row_text:
            for j, c in enumerate(cells):
                if "proyecto" in c.lower() and j + 1 < len(cells):
                    v = cells[j + 1].strip()
                    if v:
                        meta["proyecto"] = v

        if not meta.get("ubicacion") and ("ubicación" in row_text or "ubicacion" in row_text):
            for j, c in enumerate(cells):
                if "ubica" in c.lower() and j + 1 < len(cells):
                    v = cells[j + 1].strip()
                    if v:
                        meta["ubicacion"] = v

        if not meta.get("cliente") and "cliente" in row_text:
            for j, c in enumerate(cells):
                if "cliente" in c.lower() and j + 1 < len(cells):
                    v = cells[j + 1].strip()
                    if v:
                        meta["cliente"] = v

        # Latitud y longitud pueden estar en filas distintas — acumular separado
        # Solo capturar cuando la celda es exactamente "LATITUD"/"LONGITUD" (coordenada geográfica)
        # No confundir con "Longitud (m)" (longitud del perfil)
        for j, c in enumerate(cells):
            c_stripped = c.strip().lower()
            if c_stripped == "latitud" and j + 1 < len(cells):
                v = cells[j + 1].strip()
                if v:
                    lat = v
            if c_stripped == "longitud" and j + 1 < len(cells):
                v = cells[j + 1].strip()
                if v:
                    lon = v

        # Tipo perfil
        if not meta.get("tipo_perfil") and "tipo" in row_text:
            for j, c in enumerate(cells):
                if c.lower() == "tipo" and j + 1 < len(cells):
                    v = cells[j + 1].strip()
                    if v and v.lower() not in ("", "tipo"):
                        meta["tipo_perfil"] = v

        # Fecha ensayo (puede haber celdas vacías entre "Ensayo" y la fecha)
        if not meta.get("fecha_ensayo") and "ensayo" in row_text:
            for j, c in enumerate(cells):
                if "ensayo" in c.lower():
                    for k in range(j + 1, min(j + 4, len(cells))):
                        v = cells[k].strip()
                        if re.match(r"\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4}", v):
                            meta["fecha_ensayo"] = v
                            break

    if lat and lon:
        meta["coordenadas"] = f"{lat} / {lon}"

    return meta


def _parse_profundidad_table(table: list[list]) -> float | None:
    """Extrae profundidad real de la tabla 'HINCADO DIRECTO'."""
    for row in table:
        for j, c in enumerate(row):
            cs = _clean(c).lower()
            if "profundidad real" in cs and j + 1 < len(row):
                v = _num(row[j + 1])
                if v is not None:
                    return v
    return None


# ─── Formato B: tabla escalones % (Enexa) ─────────────────────────────────────

_ESCALONRE = re.compile(r"^\s*(\d{1,3})\s*%\s*$")  # require % sign to avoid false positives
_DUAL_RE   = re.compile(r"([\d,\.]+)\s*/\s*([\d,\.]+)")


def _parse_ensayo_table_b(table: list[list[str]]) -> list[dict] | None:
    """Parsea tabla de escalones % con columnas por tipo de ensayo."""
    if not table:
        return None

    # Detectar cabeceras y filas de datos
    header_rows: list[list[str]] = []
    data_rows:   list[list[str]] = []
    for row in table:
        first = (row[0] if row else "").strip()
        if _ESCALONRE.match(first):
            data_rows.append(row)
        elif not data_rows:
            header_rows.append(row)

    if not data_rows:
        return None

    # Mapear columnas → (tipo, campo)
    col_tipo: dict[int, str] = {}
    col_field: dict[int, str] = {}

    for row in header_rows:
        running = None
        for j, cell in enumerate(row):
            d = _detect_tipo(str(cell))
            if d:
                running = d
            if running and j not in col_tipo:
                col_tipo[j] = running
        for j, cell in enumerate(row):
            cl = str(cell).lower()
            if "objetivo" in cl:
                col_field[j] = "fuerza_objetivo"
            elif "aplicada" in cl:
                col_field[j] = "fuerza_aplicada"
            elif "desplazamiento" in cl or "desplaz" in cl:
                col_field[j] = "desplazamiento"

    # Si no se detectaron cabeceras, intentar asignación por posición
    if not col_field:
        max_cols = max(len(r) for r in data_rows)
        start = 2
        for ti, tipo in enumerate(TIPO_SEQUENCE):
            base = start + ti * 3
            if base + 2 < max_cols:
                col_tipo[base]     = tipo
                col_field[base]    = "fuerza_objetivo"
                col_tipo[base + 1] = tipo
                col_field[base + 1] = "fuerza_aplicada"
                col_tipo[base + 2] = tipo
                col_field[base + 2] = "desplazamiento"

    tipo_data: dict[str, dict[str, list]] = {}
    for row in data_rows:
        for j in range(len(row)):
            tipo  = col_tipo.get(j)
            field = col_field.get(j)
            if not tipo or not field:
                continue
            cell = row[j] if j < len(row) else ""
            if _is_missing(cell):
                continue
            cell = str(cell).strip()
            td = tipo_data.setdefault(tipo, {"fuerza_aplicada": [], "desplazamiento": []})
            if field == "fuerza_aplicada":
                v = _num(cell)
                if v is not None:
                    td["fuerza_aplicada"].append(v)
            elif field == "fuerza_objetivo":
                m = _DUAL_RE.search(cell)
                if m and not tipo_data.get(tipo, {}).get("fuerza_aplicada"):
                    td["fuerza_aplicada"].append(_num(m.group(2)))
            elif field == "desplazamiento":
                v = _num(cell)
                if v is not None:
                    td["desplazamiento"].append(v)

    ensayos = []
    for tipo, td in tipo_data.items():
        pairs = list(zip(td["desplazamiento"], td["fuerza_aplicada"]))
        e = _build_ensayo(tipo, pairs)
        if e:
            ensayos.append(e)
    return ensayos if ensayos else None


# ─── Extracción de metadata del texto libre ───────────────────────────────────

def _extract_metadata_text(text: str) -> dict:
    meta: dict = {}
    m = re.search(r"PROYECTO\s*[-–—]\s*([^\n\r]{3,80})", text, re.I)
    if m:
        meta["proyecto_nombre"] = m.group(1).strip()
    m = re.search(r"cliente[:\s]+([^\n\r]{3,60})", text, re.I)
    if m:
        meta["cliente"] = m.group(1).strip()
    m = re.search(r"Fecha\s+de\s+Ensayo[:\s]+([\d/\-\.]+)", text, re.I)
    if m:
        meta["fecha_ensayo"] = m.group(1).strip()
    if not meta.get("fecha_ensayo"):
        m = re.search(r"\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b", text)
        if m:
            meta["fecha_ensayo"] = m.group(1)
    m = re.search(r"Profundidad[^\d]*([\d,\.]+)\s*(mm|m\b)?", text, re.I)
    if m:
        v = _num(m.group(1))
        unit = (m.group(2) or "").lower()
        if v is not None:
            meta["profundidad_m"] = round(v / 1000, 3) if (unit == "mm" or v > 20) else v
    m = re.search(r"(?:TIPO\s+DE\s+PERFIL|Tipo\s+Perfil)[:\s]+([A-Z0-9][A-Z0-9\s/\-]*)", text, re.I)
    if m:
        meta["tipo_perfil"] = m.group(1).strip()[:30]
    if not meta.get("tipo_perfil"):
        m = re.search(r"\b(IPE\s*\d+|HEA?\s*\d+|UPN?\s*\d+|W\d+[xX]\d+|C\d+[xX\s]\d+[xX\s]\d+)\b", text, re.I)
        if m:
            meta["tipo_perfil"] = m.group(1).strip()
    m = re.search(
        r"(\d+°\s*\d+['\u2019]\s*[\d,\.]+[\"″\u201d]?\s*[NSns])"
        r"[^\n\r]{0,30}"
        r"(\d+°\s*\d+['\u2019]\s*[\d,\.]+[\"″\u201d]?\s*[EOeo])",
        text,
    )
    if m:
        meta["coordenadas"] = f"{m.group(1).strip()} / {m.group(2).strip()}"
    m = re.search(r"No\.?\s*Punto[^\n:]*[:\s]+([A-Z0-9][A-Z0-9\s\-\.]+)", text, re.I)
    if m:
        meta["punto_id"] = m.group(1).strip()[:30]
    if not meta.get("punto_id"):
        m = re.search(r"\b(POT[\s\-][A-Z0-9\-]+|PLT[\s\-]\w+)\b", text, re.I)
        if m:
            meta["punto_id"] = m.group(1).strip()
    return meta


# ─── Reconstrucción de tabla por coordenadas de palabras ──────────────────────

def _words_to_rows(words: list[dict], row_gap: float = 8.0) -> list[list[dict]]:
    if not words:
        return []
    sorted_w = sorted(words, key=lambda w: w["top"])
    rows: list[list[dict]] = []
    cur: list[dict] = [sorted_w[0]]
    cur_y = sorted_w[0]["top"]
    for w in sorted_w[1:]:
        if abs(w["top"] - cur_y) <= row_gap:
            cur.append(w)
        else:
            rows.append(sorted(cur, key=lambda x: x["x0"]))
            cur = [w]
            cur_y = w["top"]
    if cur:
        rows.append(sorted(cur, key=lambda x: x["x0"]))
    return rows


def _cluster_x(rows: list[list[dict]], merge_dist: float = 20.0) -> list[float]:
    xs: list[float] = [w["x0"] for row in rows for w in row]
    xs.sort()
    centers: list[float] = []
    for x in xs:
        for i, c in enumerate(centers):
            if abs(x - c) <= merge_dist:
                centers[i] = (c + x) / 2
                break
        else:
            centers.append(x)
    return sorted(centers)


def _rows_to_table(rows: list[list[dict]], centers: list[float]) -> list[list[str]]:
    n = len(centers)
    table: list[list[str]] = []
    for row in rows:
        cells = [""] * n
        for w in row:
            ci = min(range(n), key=lambda i: abs(centers[i] - w["x0"]))
            cells[ci] = (cells[ci] + " " + w["text"]).strip()
        table.append(cells)
    return table


# ─── Formato C: informe generado (San Martín / Enexa report) ─────────────────
#
#  Texto limpio con marcadores:
#    "Punto Analizado: POT X-N"
#    "Tipo de perfil: X | Fecha: D | Coordenadas: Y"
#    "Ensayo: Compresión Vertical (Max: N kN)"
#    tabla con columnas [Paso, Desplazamiento (mm), Fuerza (kgf), Fuerza (kN), Rigidez]

_FC_PUNTO_RE  = re.compile(r"Punto\s+Analizado:\s*(POT[\w\s\-\.\/]+?)(?:\n|$)", re.I)
_FC_ENSAYO_RE = re.compile(
    r"Ensayo:\s+(Compresi[oó]n\s+Vertical|Tensi[oó]n\s+Vertical|Tracci[oó]n\s+Vertical"
    r"|Carga\s+Lateral|Horizontal|Cortante|Compres[ií]|Tensi[oó]n|Tracci[oó]n)",
    re.I,
)
_FC_META_RE   = re.compile(
    r"Tipo\s+de\s+perfil:\s*([^|\n]+?)\s*\|[^|]*?Fecha:\s*(\S+)[^|]*?\|[^|]*?Coordenadas:\s*([^\n]+)",
    re.I,
)
_FC_ROW_RE    = re.compile(
    r"^\d+\s+([\d,\.]+|-)\s+([\d,\.]+|-)\s+([\d,\.]+|-)\s+([\d,\.]+|-)",
    re.MULTILINE,
)
_FC_HEADER_RE = re.compile(r"Paso\s+Desplazamiento", re.I)


def _fc_detect_tipo(text: str) -> str | None:
    m = _FC_ENSAYO_RE.search(text)
    if not m:
        return None
    return _detect_tipo(m.group(1))


_FC_MAX_RE = re.compile(r"\(Max:\s*([\d,\.]+)\s*kN\)", re.I)
_FC_OBS_RE = re.compile(r"Observaciones?:\s*([^\n]+)", re.I)
_FC_PROF_RE = re.compile(
    r"(?:Profundidad[^:\n]*:|hincado[^:\n]*:)\s*"
    r"([\d,\.]+)\s*(mm|m)\b",
    re.I,
)


def _fc_parse_rows(section: str) -> tuple[list[tuple[float, float]], list[float]]:
    """
    Extrae pares (desplazamiento_mm, fuerza_kg) válidos para la curva,
    y también fuerzas sin desplazamiento (fuerza_solo).
    Retorna (pairs_validos, fuerzas_solo).
    """
    pairs: list[tuple[float, float]] = []
    fuerzas_solo: list[float] = []
    # Saltar cabeceras "Paso Desplazamiento…" (puede haber más de una si corta de página)
    body = _FC_HEADER_RE.sub("", section)
    for m in _FC_ROW_RE.finditer(body):
        disp_s, kgf_s, kn_s = m.group(1), m.group(2), m.group(3)
        fuerza: float | None = None
        if kgf_s != "-":
            fuerza = _num(kgf_s)
        elif kn_s != "-":
            v = _num(kn_s)
            if v is not None:
                fuerza = v / KGF_TO_KN
        if fuerza is None or fuerza <= 0:
            continue
        disp = _num(disp_s) if disp_s != "-" else None
        if disp is not None:
            pairs.append((disp, fuerza))
        else:
            fuerzas_solo.append(fuerza)
    return pairs, fuerzas_solo


def _build_ensayo_fc(tipo: str, data_text: str) -> dict | None:
    """
    Construye un ensayo desde la sección de texto de Format C.
    Incluye curva si hay datos de desplazamiento; si no, reporta max de texto.
    """
    pairs, fuerzas_solo = _fc_parse_rows(data_text)
    max_kn_m = _FC_MAX_RE.search(data_text)
    max_kn_text = _num(max_kn_m.group(1)) if max_kn_m else None

    if pairs:
        e = _build_ensayo(tipo, pairs)
        if e and max_kn_text:
            # Si el max del texto es mayor que lo calculado, usarlo
            max_kgf_text = max_kn_text / KGF_TO_KN
            if max_kgf_text > e["carga_maxima_kgf"]:
                e["carga_maxima_kgf"] = round(max_kgf_text, 2)
        return e

    # Sin desplazamiento pero con fuerza — registrar max force
    max_kgf = max(fuerzas_solo) if fuerzas_solo else None
    if max_kgf is None and max_kn_text:
        max_kgf = max_kn_text / KGF_TO_KN
    if max_kgf is None:
        return None

    return {
        "tipo": tipo,
        "nombre": TIPO_LABELS.get(tipo, tipo),
        "carga_maxima_kgf": round(max_kgf, 2),
        "desplazamiento_maximo_mm": None,
        "puntos": [],
        "sin_datos_desplazamiento": True,
    }


def _parse_format_c(full_text: str) -> dict | None:
    """Parsea el formato San Martín / informe generado desde texto plano."""
    if not _FC_PUNTO_RE.search(full_text):
        return None  # No es formato C

    proyecto: dict = {
        "nombre": "Proyecto POT", "cliente": "", "ubicacion": "",
        "fecha": str(date.today()),
    }

    # Extraer nombre de proyecto del encabezado (antes del primer "Punto Analizado")
    header = _FC_PUNTO_RE.split(full_text)[0]
    # Busca línea 2 (nombre del proyecto) y línea 3 (cliente · ubicación · fecha)
    lines = [l.strip() for l in header.splitlines() if l.strip()]
    for i, line in enumerate(lines):
        # "INFORME POT …" / "SAN MARTIN CESAR" / "Cliente · Lugar · Fecha"
        if "informe pot" in line.lower() or "pull out test" in line.lower():
            if i + 1 < len(lines):
                proyecto["nombre"] = lines[i + 1]
            if i + 2 < len(lines):
                parts = re.split(r"\s*·\s*", lines[i + 2])
                if len(parts) >= 1:
                    proyecto["cliente"] = parts[0].strip()
                if len(parts) >= 2:
                    proyecto["ubicacion"] = parts[1].strip()
                if len(parts) >= 3:
                    proyecto["fecha"] = parts[2].strip()
            break

    puntos_by_id: dict[str, dict] = {}
    debug_pages: list[dict] = []

    # Dividir por "Punto Analizado:" — cada chunk es un punto
    chunks = _FC_PUNTO_RE.split(full_text)
    # chunks = [header, punto_id_1, body_1, punto_id_2, body_2, …]
    i = 1
    while i < len(chunks) - 1:
        raw_pid = chunks[i].strip()
        body    = chunks[i + 1]
        i += 2

        punto_id = re.sub(r"\s+", "-", raw_pid.strip())

        # Metadata
        meta_m = _FC_META_RE.search(body)
        tipo_perfil  = meta_m.group(1).strip() if meta_m else None
        fecha_ensayo = meta_m.group(2).strip() if meta_m else None
        coordenadas  = meta_m.group(3).strip() if meta_m else None

        # Profundidad
        prof = None
        pm = _FC_PROF_RE.search(body)
        if pm:
            v = _num(pm.group(1))
            unit = pm.group(2).lower()
            if v is not None:
                prof = round(v / 1000, 3) if unit == "mm" else v

        # Observaciones
        obs_m = _FC_OBS_RE.search(body)
        observaciones = obs_m.group(1).strip() if obs_m else None

        if punto_id not in puntos_by_id:
            p = _empty_punto(punto_id)
            p["tipo_perfil"]    = tipo_perfil
            p["coordenadas"]    = coordenadas
            p["fecha_ensayo"]   = fecha_ensayo
            p["profundidad_m"]  = prof
            p["observaciones"]  = observaciones
            puntos_by_id[punto_id] = p

        # Dividir por "Ensayo: X" para obtener secciones de ensayo
        ensayo_parts = _FC_ENSAYO_RE.split(body)
        # ensayo_parts = [pre_text, tipo_text_1, data_1, tipo_text_2, data_2, …]
        j = 1
        ensayos_found = []
        while j < len(ensayo_parts) - 1:
            tipo_text = ensayo_parts[j]
            data_text = ensayo_parts[j + 1]
            j += 2

            tipo = _detect_tipo(tipo_text)
            if tipo is None:
                continue

            e = _build_ensayo_fc(tipo, data_text)
            if e:
                ensayos_found.append(tipo)
                existing = {en["tipo"] for en in puntos_by_id[punto_id]["ensayos"]}
                if tipo not in existing:
                    puntos_by_id[punto_id]["ensayos"].append(e)

        if ensayos_found:
            debug_pages.append({
                "punto_id": punto_id,
                "ensayos": ensayos_found,
                "formato": "C",
            })

    if not puntos_by_id:
        return None

    return {
        "proyecto": proyecto,
        "puntos":   list(puntos_by_id.values()),
        "_pdf_debug": debug_pages,
    }


# ─── Formato D: Enexa "FORMATO PARA PRUEBA PULL OUT TEST" texto espaciado ─────
#
#  Una página por punto con encabezado espaciado "F O R M A T O P A R A ...".
#  Columnas: Escalón% | Tiempo | Comp(Obj/Aplic/Desp) | Tract(Obj/Aplic/Desp) | Lat(Obj/Aplic/Desp)
#  Punto ID al final de página: "P 0 T A -1 - S A T"

_FD_DETECT_RE = re.compile(
    r"F\s+O\s+R\s+M\s+A\s+T\s+O\s+P\s+A\s+R\s+A|FORMATO\s+PARA\s+PRUEBA",
    re.I,
)

# Fila de datos: "0% 30 seg 1,5/146,8 153 0 1,2/117,44 - - 1,4/139,2 149 0,19"
# Grupos: 1=escalon, 2=comp_obj, 3=comp_aplic, 4=comp_disp,
#         5=tract_obj, 6=tract_aplic, 7=tract_disp,
#         8=lat_obj,  9=lat_aplic, 10=lat_disp
_FD_ROW_RE = re.compile(
    r"^(\d{1,3})\s*%\s+"
    r"[\d,\.]+\s+(?:seg|min|s)\s+"
    r"([\d,\.]+/[\d,\.]+|-)\s+([\d,\.]+|-)\s+([\d,\.]+|-)\s+"
    r"([\d,\.]+/[\d,\.]+|-)\s+([\d,\.]+|-)\s+([\d,\.]+|-)\s+"
    r"([\d,\.]+/[\d,\.]+|-)\s+([\d,\.]+|-)\s+([\d,\.]+|-)",
    re.MULTILINE,
)


def _parse_format_d(pages_text: list[str]) -> dict | None:
    """
    Parse Format D: Enexa 'FORMATO PARA PRUEBA PULL OUT TEST'.
    Una página por punto. Encabezado espaciado 'F O R M A T O P A R A ...'.
    """
    if not any(_FD_DETECT_RE.search(pt) for pt in pages_text):
        return None

    proyecto: dict = {
        "nombre": "Proyecto POT", "cliente": "", "ubicacion": "",
        "fecha": str(date.today()),
    }
    puntos_by_id: dict[str, dict] = {}
    debug_pages: list[dict] = []

    for page_num, raw_text in enumerate(pages_text):
        if not _FD_DETECT_RE.search(raw_text):
            continue

        # Eliminar separadores | que pdfplumber puede o no incluir
        text = raw_text.replace("|", " ")

        # ── Filas de escalón ──────────────────────────────────────────────
        # Secuencia fija de tipos: Compresión, Tracción, Lateral
        tipo_aplics: dict[str, list] = {
            "compresion_vertical": [],
            "tension_vertical":    [],
            "carga_lateral":       [],
        }
        tipo_disps: dict[str, list] = {
            "compresion_vertical": [],
            "tension_vertical":    [],
            "carga_lateral":       [],
        }

        for m in _FD_ROW_RE.finditer(text):
            for tipo, ai, di in [
                ("compresion_vertical", 3, 4),
                ("tension_vertical",    6, 7),
                ("carga_lateral",       9, 10),
            ]:
                aplic_s = m.group(ai)
                disp_s  = m.group(di)
                a = _num(aplic_s) if aplic_s != "-" else None
                d = _num(disp_s)  if disp_s  != "-" else None
                if a is not None and a > 0:
                    tipo_aplics[tipo].append(a)
                    tipo_disps[tipo].append(d if d is not None else 0.0)

        # ── Construir ensayos ─────────────────────────────────────────────
        ensayos: list[dict] = []
        for tipo in ("compresion_vertical", "tension_vertical", "carga_lateral"):
            aplics = tipo_aplics[tipo]
            disps  = tipo_disps[tipo]
            if not aplics:
                continue
            pairs = list(zip(disps, aplics))
            e = _build_ensayo(tipo, pairs)
            if e:
                ensayos.append(e)

        if not ensayos:
            continue

        # ── Metadata ──────────────────────────────────────────────────────
        meta = _extract_metadata_text(text)

        # Nombre de proyecto desde "PROYECTO - SAN MARTIN CESAR"
        pm = re.search(r"PROYECTO\s*[-–]\s*([^\n\r]{3,60})", text, re.I)
        if pm and proyecto["nombre"] == "Proyecto POT":
            proyecto["nombre"] = pm.group(1).strip()

        # Fecha global del proyecto desde encabezado "Fecha: DD/MM/YYYY"
        if proyecto["fecha"] == str(date.today()):
            fm = re.search(r"Fecha:\s*([\d/\-\.]{6,10})", text, re.I)
            if fm:
                proyecto["fecha"] = fm.group(1).strip()

        # Punto ID desde "P 0 T A -1 - S A T   Coordenadas"
        punto_id = meta.get("punto_id")
        if not punto_id:
            pid_m = re.search(
                r"P\s+[0O]\s+T\s+((?:[A-Z0-9\-\s]+?))(?:\s{3,}|Coordenadas|\n)",
                text, re.I,
            )
            if pid_m:
                raw_suffix = pid_m.group(1).strip()
                # Colapsar espacios: "A -1 - S A T" → "A-1-SAT"
                punto_id = "POT-" + re.sub(r"\s+", "", raw_suffix).upper()
            else:
                punto_id = f"P-{page_num + 1:02d}"

        if punto_id not in puntos_by_id:
            p = _empty_punto(punto_id)
            p["profundidad_m"] = meta.get("profundidad_m")
            p["tipo_perfil"]   = meta.get("tipo_perfil")
            p["coordenadas"]   = meta.get("coordenadas")
            p["fecha_ensayo"]  = meta.get("fecha_ensayo")
            puntos_by_id[punto_id] = p

        existing = {e["tipo"] for e in puntos_by_id[punto_id]["ensayos"]}
        for e in ensayos:
            if e["tipo"] not in existing:
                puntos_by_id[punto_id]["ensayos"].append(e)
                existing.add(e["tipo"])

        debug_pages.append({
            "page": page_num + 1,
            "punto_id": punto_id,
            "ensayos": [e["tipo"] for e in ensayos],
            "formato": "D",
        })

    if not puntos_by_id:
        return None

    return {
        "proyecto": proyecto,
        "puntos":   list(puntos_by_id.values()),
        "_pdf_debug": debug_pages,
    }


# ─── Formato E: GLO-HAT-POT (GLOBALEM "FORMATO RECOPILACIÓN DE DATOS") ────────
#
# Estructura por página:
#   Cabecera: "PULL OUT TEST - FORMATO RECOPILACIÓN DE DATOS"
#   Sección 1: ID Punto, Perfil, L[m], L_amp[m], Coordenadas, Tipo de Prueba
#   Sección 3: tabla  Etapa | Nominal[Kg][kN] | Real[Kg][kN] | δ[mm]
#
# Cada página = un punto con un solo ensayo.

_GLO_SIGNAL_RE = re.compile(r"FORMATO\s+RECOPILACI[OÓ]N\s+DE\s+DATOS", re.I)


def _parse_format_e(pages_text: list[str], file_bytes: bytes) -> dict | None:
    import pdfplumber

    if not any(_GLO_SIGNAL_RE.search(t) for t in pages_text):
        return None

    proyecto: dict = {}
    puntos: list[dict] = []
    debug_pages: list[dict] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for i, (page_text, page) in enumerate(zip(pages_text, pdf.pages)):
            if not _GLO_SIGNAL_RE.search(page_text):
                continue

            # ── Metadata ────────────────────────────────────────────────────
            punto_id = None
            m = re.search(r"ID\s+Punto\s+(\S+)", page_text)
            if m:
                punto_id = "ID-" + m.group(1).strip()
            if not punto_id:
                continue

            tipo_perfil = None
            m = re.search(r"Perfil\s+([CW]\s*[\w\s\d\.\(\)x]+?)(?:\s{2,}|Coordenadas|\n)", page_text, re.I)
            if m:
                tipo_perfil = m.group(1).strip()

            profundidad_m = None
            # La tabla de metadata extrae "L [m]\nemp" como etiqueta con valor en columna adyacente.
            # En el texto plano aparece como segundo "L [m]" en la línea de "Rechazo No L [m] X.XX"
            m = re.search(r"Rechazo\s+(?:No|S[ií])\s+([\d\.]+)", page_text, re.I)
            if m:
                profundidad_m = _num(m.group(1))

            coordenadas = None
            m = re.search(r"Coordenadas\s+([\d°'\"\.NS\s]+)\n\s*([\d°'\"\.OEW\s]+)", page_text)
            if m:
                coordenadas = m.group(1).strip() + " / " + m.group(2).strip()

            fecha_ensayo = None
            m = re.search(r"Fecha\s+([\d/]+)", page_text)
            if m:
                fecha_ensayo = m.group(1).strip()

            # Datos del proyecto (primera página con datos)
            if not proyecto.get("nombre"):
                m = re.search(r"Proyecto\s+(.+?)(?:\n|Ubicaci)", page_text, re.I)
                if m:
                    proyecto["nombre"] = m.group(1).strip()
                m = re.search(r"Ubicaci[oó]n\s+(.+?)(?:\s{2,}|Fecha|\n)", page_text, re.I)
                if m:
                    proyecto["ubicacion"] = m.group(1).strip()
                m = re.search(r"Cliente\s+(.+?)(?:\s{2,}|Hora|\n)", page_text, re.I)
                if m:
                    proyecto["cliente"] = m.group(1).strip()
                if fecha_ensayo:
                    proyecto["fecha"] = fecha_ensayo

            # ── Tipo de ensayo ───────────────────────────────────────────────
            tipo = None
            m = re.search(r"3\.\s*RESULTADOS\s+DE\s+PRUEBA\s*[-–]\s*(.+?)(?:\.|$)", page_text, re.I | re.M)
            if m:
                tipo = _detect_tipo(m.group(1))
            if tipo is None:
                m = re.search(r"Tipo\s+de\s+Prueba\s+(.+?)(?:\n|$)", page_text, re.I)
                if m:
                    tipo = _detect_tipo(m.group(1))
            if tipo is None:
                continue

            # ── Tabla de resultados ──────────────────────────────────────────
            pairs: list[tuple] = []

            for table in page.extract_tables():
                if not table:
                    continue
                # Detectar tabla de resultados: debe tener "Etapa" en las primeras filas
                header_text = " ".join(
                    str(c or "") for row in table[:4] for c in row
                ).lower()
                if "etapa" not in header_text:
                    continue

                # La estructura es siempre:
                #   col 0: Etapa
                #   col 1: Nominal [Kg]   col 2: Nominal [kN]
                #   col 3: Real [Kg]      col 4: Real [kN]
                #   col 5: δ [mm]
                # Buscamos el índice de δ dinámicamente (última columna con "mm" o "δ")
                col_disp = None
                for hi in range(min(4, len(table))):
                    for j, cell in enumerate(table[hi]):
                        cs = str(cell or "").lower()
                        if "δ" in cs or ("\u03b4" in cs) or (
                            "mm" in cs and "kg" not in cs and "kn" not in cs and "nominal" not in cs
                        ):
                            col_disp = j

                # Fallback: última columna
                if col_disp is None and table:
                    col_disp = len(table[0]) - 1

                # Real [Kg] = col_disp - 2
                col_kg_real = col_disp - 2 if col_disp is not None and col_disp >= 2 else None

                for row in table:
                    if not row:
                        continue
                    etapa = str(row[0] or "").strip().lower()
                    # Saltar cabeceras y fila de Rotura (representación sintética del fallo)
                    # NOTA: no incluir "" en skip_words porque "" in cualquier_str es siempre True
                    skip_words = ("etapa", "carga", "nominal", "real", "kg", "kn", "[", "rotura")
                    if not etapa or any(sw in etapa for sw in skip_words):
                        continue

                    dv = row[col_disp] if col_disp is not None and col_disp < len(row) else None
                    fv = row[col_kg_real] if col_kg_real is not None and col_kg_real < len(row) else None

                    d = _num(dv)
                    f = _num(fv)
                    if d is not None and f is not None and f > 0:
                        pairs.append((d, f))

                if pairs:
                    break

            if not pairs:
                continue

            ensayo = _build_ensayo(tipo, pairs)
            if not ensayo:
                continue

            punto = _empty_punto(punto_id)
            punto["profundidad_m"] = profundidad_m
            punto["tipo_perfil"] = tipo_perfil
            punto["coordenadas"] = coordenadas
            punto["fecha_ensayo"] = fecha_ensayo
            punto["ensayos"] = [ensayo]
            puntos.append(punto)
            debug_pages.append({
                "page": i + 1, "punto_id": punto_id,
                "ensayos": [tipo], "formato": "E",
            })

    if not puntos:
        return None

    if not proyecto.get("nombre"):
        proyecto["nombre"] = "Proyecto POT"

    return {"proyecto": proyecto, "puntos": puntos, "_pdf_debug": debug_pages}


# ─── Parser PDF principal ─────────────────────────────────────────────────────

def parse_pdf(file_bytes: bytes) -> dict:
    import pdfplumber

    # Extraer texto de cada página una sola vez (reutilizado por varios parsers)
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        pages_text = [
            page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            for page in pdf.pages
        ]
    all_text = "\n".join(pages_text)

    # ── Intento 1: Formato E (GLO-HAT-POT "FORMATO RECOPILACIÓN DE DATOS") ─────
    result_e = _parse_format_e(pages_text, file_bytes)
    if result_e and result_e.get("puntos"):
        return result_e

    # ── Intento 2: Formato C (texto limpio, "Punto Analizado:") ───────────────
    result_c = _parse_format_c(all_text)
    if result_c and result_c.get("puntos"):
        return result_c

    # ── Intento 3: Formato D (Enexa spaced-text "FORMATO PARA PRUEBA") ───────
    result_d = _parse_format_d(pages_text)
    if result_d and result_d.get("puntos"):
        return result_d

    # ── Intento 3: Formato A/B (tablas pdfplumber) ────────────────────────────
    proyecto: dict = {
        "nombre": "Proyecto POT", "cliente": "", "ubicacion": "",
        "fecha": str(date.today()),
    }
    puntos_by_id: dict[str, dict] = {}  # punto_id → punto dict
    debug_pages: list[dict] = []

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_num, page in enumerate(pdf.pages):
            full_text = pages_text[page_num]
            text_meta = _extract_metadata_text(full_text)

            # Actualizar proyecto con metadata global (texto plano)
            if text_meta.get("proyecto_nombre") and proyecto["nombre"] == "Proyecto POT":
                proyecto["nombre"] = text_meta["proyecto_nombre"]
            if text_meta.get("cliente") and not proyecto["cliente"]:
                proyecto["cliente"] = text_meta["cliente"]
            if text_meta.get("fecha_ensayo") and proyecto["fecha"] == str(date.today()):
                proyecto["fecha"] = text_meta["fecha_ensayo"]

            # ── Extraer tablas con pdfplumber ─────────────────────────────
            tables = page.extract_tables()

            # ── Intentar Formato A (Pivijay): tablas individuales ─────────
            page_meta: dict = {}
            ensayos_a: list[dict] = []
            profundidad: float | None = None

            for tbl in tables:
                tbl_str = [[_clean(c) for c in row] for row in tbl]
                if not tbl_str or not tbl_str[0]:
                    continue

                title_raw = " ".join(_decode_repeated(c) for c in tbl_str[0] if c)

                # ¿Es tabla de ensayo?
                if _ENSAYO_TITLE_RE.search(title_raw):
                    tipo, pts = _parse_ensayo_table_a(tbl_str)
                    if tipo and pts:
                        e = {
                            "tipo": tipo,
                            "nombre": TIPO_LABELS.get(tipo, tipo),
                            "carga_maxima_kgf":         max(p["fuerza_kg"]         for p in pts),
                            "desplazamiento_maximo_mm": max(p["desplazamiento_mm"] for p in pts),
                            "puntos": pts,
                        }
                        ensayos_a.append(e)
                    continue

                # ¿Es tabla de metadata del punto?
                # Revisar las primeras 3 filas (la fila 0 puede ser el título "REPORTE DE ENSAYO PULL OUT TEST")
                first_rows_decoded = " ".join(
                    _decode_repeated(c)
                    for row in tbl_str[:3]
                    for c in row
                    if c
                )
                if any(k in first_rows_decoded.upper() for k in ("PROYECTO", "CLIENTE", "UBICAC", "POT", "CORDENADA")):
                    m = _parse_metadata_table_a(tbl_str)
                    page_meta.update({k: v for k, v in m.items() if v})
                    continue

                # ¿Es tabla de profundidad?
                flat = " ".join(c for row in tbl_str for c in row).lower()
                if "hincado directo" in flat or "profundidad real" in flat:
                    p = _parse_profundidad_table(tbl_str)
                    if p is not None:
                        profundidad = p
                    continue

            # ── Si no hay ensayos Formato A, intentar Formato B ───────────
            ensayos_b: list[dict] = []
            if not ensayos_a:
                # Intentar reconstrucción por palabras
                words = page.extract_words(x_tolerance=3, y_tolerance=3, keep_blank_chars=False)
                if words:
                    rows    = _words_to_rows(words, row_gap=8)
                    centers = _cluster_x(rows, merge_dist=20)
                    table   = _rows_to_table(rows, centers)
                    result  = _parse_ensayo_table_b(table)
                    if result:
                        ensayos_b = result

                # Fallback: extract_tables con estrategia de texto
                if not ensayos_b:
                    for strat in [
                        {"vertical_strategy": "lines",  "horizontal_strategy": "lines"},
                        {"vertical_strategy": "text",   "horizontal_strategy": "text"},
                    ]:
                        try:
                            for tbl in page.extract_tables(strat):
                                tbl_str = [[_clean(c) for c in row] for row in tbl]
                                r = _parse_ensayo_table_b(tbl_str)
                                if r:
                                    ensayos_b = r
                                    break
                        except Exception:
                            pass
                        if ensayos_b:
                            break

            ensayos = ensayos_a or ensayos_b
            if not ensayos:
                continue

            # Propagar nombre de proyecto / cliente / ubicación / fecha desde page_meta
            if page_meta.get("proyecto") and proyecto["nombre"] == "Proyecto POT":
                proyecto["nombre"] = page_meta["proyecto"]
            if page_meta.get("cliente") and not proyecto["cliente"]:
                proyecto["cliente"] = page_meta["cliente"]
            if page_meta.get("ubicacion") and not proyecto["ubicacion"]:
                proyecto["ubicacion"] = page_meta["ubicacion"]
            if page_meta.get("fecha_ensayo") and proyecto["fecha"] == str(date.today()):
                proyecto["fecha"] = page_meta["fecha_ensayo"]

            # Determinar punto_id
            punto_id = (
                page_meta.get("punto_id")
                or text_meta.get("punto_id")
                or f"P-{page_num + 1:02d}"
            )

            # Agrupar o crear el punto
            if punto_id not in puntos_by_id:
                p = _empty_punto(punto_id)
                p["profundidad_m"] = profundidad or text_meta.get("profundidad_m")
                p["tipo_perfil"]   = page_meta.get("tipo_perfil") or text_meta.get("tipo_perfil")
                p["coordenadas"]   = page_meta.get("coordenadas") or text_meta.get("coordenadas")
                p["fecha_ensayo"]  = page_meta.get("fecha_ensayo") or text_meta.get("fecha_ensayo")
                puntos_by_id[punto_id] = p

            # Agregar ensayos (sin duplicar por tipo)
            existing_tipos = {e["tipo"] for e in puntos_by_id[punto_id]["ensayos"]}
            for e in ensayos:
                if e["tipo"] not in existing_tipos:
                    puntos_by_id[punto_id]["ensayos"].append(e)
                    existing_tipos.add(e["tipo"])

            debug_pages.append({
                "page": page_num + 1,
                "punto_id": punto_id,
                "ensayos": [e["tipo"] for e in ensayos],
                "formato": "A" if ensayos_a else "B",
            })

    puntos = list(puntos_by_id.values())
    return {
        "proyecto": proyecto,
        "puntos":   puntos,
        "_pdf_debug": debug_pages,
    }


# ─── XLSX parser ──────────────────────────────────────────────────────────────

def _find_header_row(rows: list) -> int | None:
    for i, row in enumerate(rows):
        cells = [str(c or "").lower() for c in row]
        if any("desplazamiento" in c for c in cells) and \
           any("fuerza" in c or ("kg" in c and "kg/m" not in c) for c in cells):
            return i
    return None


def _map_xlsx_columns(rows: list, hi: int) -> dict[int, tuple[str, str]]:
    header = rows[hi]
    tipo_by_col: dict[int, str] = {}
    for lb in range(1, min(5, hi + 1)):
        running = None
        for j, cell in enumerate(rows[hi - lb]):
            if cell is not None:
                d = _detect_tipo(str(cell))
                if d:
                    running = d
            if running and j not in tipo_by_col:
                tipo_by_col[j] = running
    col_map: dict[int, tuple[str, str]] = {}
    dc = 0
    for j, cell in enumerate(header):
        cs = str(cell or "").lower().strip()
        if "desplazamiento" in cs:
            tipo = tipo_by_col.get(j) or (TIPO_SEQUENCE[dc] if dc < len(TIPO_SEQUENCE) else f"t{dc}")
            col_map[j] = (tipo, "disp")
            dc += 1
        elif "fuerza" in cs or ("kg" in cs and "kg/m" not in cs):
            near = max((k for k, v in col_map.items() if v[1] == "disp" and k < j), default=None)
            tipo = col_map[near][0] if near is not None else tipo_by_col.get(j, "tension_vertical")
            col_map[j] = (tipo, "fuerza")
    return col_map


def _extract_xlsx_ensayos(data_rows: list, tipo_cols: dict) -> list[dict]:
    ensayos = []
    for tipo, cols in tipo_cols.items():
        if "disp" not in cols or "fuerza" not in cols:
            continue
        di, fi = cols["disp"], cols["fuerza"]
        pairs = [
            (row[di], row[fi])
            for row in data_rows
            if row and len(row) > max(di, fi)
            and not _is_missing(row[di]) and not _is_missing(row[fi])
        ]
        e = _build_ensayo(tipo, pairs)
        if e:
            ensayos.append(e)
    return ensayos


_MARKER_RE = re.compile(r"^(ensayo|plt|pot|punto|p)[^\w]?\s*(\d+)", re.IGNORECASE)


def parse_xlsx(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    puntos: list[dict] = []
    pnombre = "Proyecto POT"
    pfecha  = str(date.today())

    for sheet_name in wb.sheetnames:
        ws   = wb[sheet_name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        for row in rows[:6]:
            for cell in row:
                cs = str(cell or "")
                m = re.search(r"PROYECTO\s*[-–]\s*(.+)", cs, re.I)
                if m:
                    pnombre = m.group(1).strip()[:80]
                if re.match(r"\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4}", cs):
                    pfecha = cs

        hi = _find_header_row(rows)
        if hi is None:
            continue
        col_map   = _map_xlsx_columns(rows, hi)
        tipo_cols: dict[str, dict[str, int]] = {}
        for idx, (tipo, field) in col_map.items():
            tipo_cols.setdefault(tipo, {})[field] = idx
        if not tipo_cols:
            continue

        data_rows = rows[hi + 1:]
        markers = [
            (i, row[0])
            for i, row in enumerate(data_rows)
            if row and row[0] is not None and _MARKER_RE.match(str(row[0]).strip())
        ]
        if markers:
            bounds = markers + [(len(data_rows), None)]
            for k, (start, label) in enumerate(markers):
                end = bounds[k + 1][0]
                mm  = _MARKER_RE.match(str(label).strip())
                pid = f"P-{mm.group(2).zfill(2)}" if mm else str(label)[:20]
                ens = _extract_xlsx_ensayos(data_rows[start + 1: end], tipo_cols)
                if ens:
                    p = _empty_punto(pid)
                    p["ensayos"] = ens
                    puntos.append(p)
        else:
            pid = re.sub(r"[^\w\-]", "-", sheet_name.strip())[:20] or "P-01"
            ens = _extract_xlsx_ensayos(data_rows, tipo_cols)
            if ens:
                p = _empty_punto(pid)
                p["ensayos"] = ens
                puntos.append(p)

    return {
        "proyecto": {"nombre": pnombre, "cliente": "", "ubicacion": "", "fecha": pfecha},
        "puntos":   puntos,
    }


# ─── Parser texto plano ───────────────────────────────────────────────────────

_TIPO_SPLIT_RE = re.compile(
    r"(tensión\s+vertical|tension\s+vertical|compresión\s+vertical|compresion\s+vertical"
    r"|carga\s+lateral|tracción|traccion|prueba\s+de\s+compres[ió]n"
    r"|ensayo\s+de\s+carga\s+horizontal|ensayo\s+de\s+carga\s+vertical)",
    re.IGNORECASE,
)
_PUNTO_SPLIT_RE = re.compile(r"(?:PLT|POT|Ensayo|Punto)[\s\-]*(\d+)", re.IGNORECASE)
_ROW_RE = re.compile(r"^\s*([+-]?\d+(?:[,\.]\d+)?)\s+([+-]?\d+(?:[,\.]\d+)?)\s*$", re.MULTILINE)


def _text_proyecto(text: str) -> dict:
    p = {"nombre": "Proyecto POT", "cliente": "", "ubicacion": "", "fecha": str(date.today())}
    m = re.search(r"PROYECTO\s*[-–]\s*([^\n\r]+)", text, re.I)
    if m:
        p["nombre"] = m.group(1).strip()[:80]
    m = re.search(r"\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b", text)
    if m:
        p["fecha"] = m.group(1)
    return p


def parse_text(text: str) -> dict:
    proyecto = _text_proyecto(text)
    puntos: list[dict] = []
    parts = _PUNTO_SPLIT_RE.split(text)
    if len(parts) >= 3:
        i = 1
        while i < len(parts) - 1:
            content = parts[i + 1]
            subs = _TIPO_SPLIT_RE.split(content)
            ensayos = []
            j = 1
            while j < len(subs) - 1:
                tipo = _detect_tipo(subs[j])
                if tipo:
                    pairs = _ROW_RE.findall(subs[j + 1])
                    e = _build_ensayo(tipo, [(a, b) for a, b in pairs])
                    if e:
                        ensayos.append(e)
                j += 2
            if ensayos:
                p = _empty_punto(f"P-{parts[i].strip().zfill(2)}")
                p["ensayos"] = ensayos
                puntos.append(p)
            i += 2
    return {"proyecto": proyecto, "puntos": puntos}


# ─── Entrada principal ────────────────────────────────────────────────────────

def analyze_file(file_bytes: bytes, filename: str) -> dict:
    ext = filename.lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        data = parse_xlsx(file_bytes)
        note = "Parsed with openpyxl."
    elif ext == "pdf":
        data = parse_pdf(file_bytes)
        note = "Parsed with pdfplumber (no AI). Soporta formatos Pivijay (A), Enexa escalon (B), informe generado (C), Enexa FORMATO PARA PRUEBA (D) y GLO-HAT-POT GLOBALEM (E)."
    elif ext == "txt":
        data = parse_text(file_bytes.decode("utf-8", errors="replace"))
        note = "Parsed as plain text."
    else:
        raise ValueError(f"Formato no soportado: .{ext}. Usa PDF, XLSX o TXT.")

    pdf_debug = data.pop("_pdf_debug", None)
    data["_debug"] = {
        "filename": filename, "ext": ext,
        "parser": "direct (no AI)",
        "note": note,
        "puntos_found": len(data.get("puntos", [])),
        **({"pdf_pages_with_data": pdf_debug} if pdf_debug else {}),
    }
    return _enrich(data)


# ─── Enriquecimiento ──────────────────────────────────────────────────────────

DISP_SATISFACTORIO = 15.0
DISP_REDISENO      = 25.4


def _criterio_ensayo(disp_max: float, tipo: str = "") -> tuple[bool, str]:
    if disp_max <= DISP_SATISFACTORIO:
        return True, "satisfactorio"
    if tipo == "carga_lateral":
        return False, "requiere_rediseno"
    if disp_max <= DISP_REDISENO:
        return False, "no_cumple_deformaciones"
    return False, "requiere_rediseno"


def _enrich(data: dict) -> dict:
    for punto in data.get("puntos", []):
        punto_ok = True
        peor  = "satisfactorio"
        orden = ["satisfactorio", "no_cumple_deformaciones", "requiere_rediseno"]
        for e in punto.get("ensayos", []):
            kgf_max  = e.get("carga_maxima_kgf") or 0
            disp_max = e.get("desplazamiento_maximo_mm") or 0
            e["carga_maxima_kn"] = round(kgf_max * KGF_TO_KN, 3)
            # Ensayos sin datos de desplazamiento → no evaluado
            if e.get("sin_datos_desplazamiento"):
                e["cumple_criterio"] = None
                e["estado_criterio"] = "no_evaluado"
            else:
                cumple, estado = _criterio_ensayo(disp_max, e.get("tipo", ""))
                e["cumple_criterio"] = cumple
                e["estado_criterio"] = estado
                if not cumple:
                    punto_ok = False
                if orden.index(estado) > orden.index(peor):
                    peor = estado
            for pt in e.get("puntos", []):
                kg   = pt.get("fuerza_kg") or 0
                disp = pt.get("desplazamiento_mm") or 0
                pt["fuerza_kn"]     = round(kg * KGF_TO_KN, 4)
                pt["rigidez_kn_mm"] = round(pt["fuerza_kn"] / disp, 4) if disp else None
        punto["cumple_criterio"] = punto_ok
        punto["estado"]          = peor
    return data
